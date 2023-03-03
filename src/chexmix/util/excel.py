from typing import Dict

import export_data as export_data
import pandas as pd
from importlib_resources import files
from openpyxl import Workbook


def write_excel(ws, tax_data: Dict[str, Dict], count_tax: Dict[str, int] = {}, row: int = 2, col: int = 1) -> int:
    '''
    Write hierarchical data to excel.
    ---example---
    |TAX_ID|SCIENTIFIC_NAME|Unnamed: 2|Unnamed: 3
    |58880 | Phyllanthus   |296047    |Phyllanthus ussuriensis
    |      |               |          |
    |2706  | Citrus        |55188     | Citrus unshiu
    |      |               |135197    | Citrus junos

    :param ws: worksheet
    :param tax_data: hierarchical dictionary data to be extracted with excel
    :param count_tax: data with the number of appearances
    :param row: row
    :param col: col
    :return: next rows
    '''
    for node in tax_data:
        ws.cell(row, col, node)
        ws.cell(row, col + 1, tax_data[node]['name'])
        if len(count_tax) != 0:
            ws.cell(row, col + 2, count_tax[node])
            row = write_excel(ws, tax_data[node]['children'], count_tax, row, col+3)
        else:
            row = write_excel(ws, tax_data[node]['children'], count_tax, row, col+2)
        row += 1
    return row


def xlsx_to_csv(path: str):
    '''
    :param path: file path
    :return: None
    '''
    xlsx = pd.read_excel(path)
    xlsx.to_csv(f"{path.split('.')[0]}.csv", header=False, index=False, float_format='%.f')


def dict_to_excel(file_name: str, tax_data: Dict[str, Dict], count_tax: Dict[str, int] = {}):
    '''
    Define column name and write hierarchical data to excel by calling reculsive func.
    And convert xlsx file to csv file

    :param input_name: keywords (filename)
    :param tax_data: hierarchical dictionary data to be extracted with excel
    :param count_tax: data with the number of appearances
    :return: None
    '''
    wb = Workbook()
    ws = wb.active
    file_path = str(files(export_data).joinpath(f'{file_name}.xlsx'))

    ws.cell(1, 1, 'TAX_ID')
    ws.cell(1, 2, 'SCIENTIFIC_NAME')
    if len(count_tax) != 0:
        ws.cell(1, 3, 'COUNT')

    write_excel(ws, tax_data, count_tax)
    wb.save(file_path)
    xlsx = pd.read_excel(file_path)
    xlsx.to_csv(f"{file_path.split('.')[0]}.csv", header=True, index=False, float_format='%.f')
